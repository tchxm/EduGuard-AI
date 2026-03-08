import os
import json
from datetime import datetime
from flask import render_template, redirect, url_for, request, flash, jsonify, send_file
from flask_login import login_required, login_user, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import face_recognition
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import smtplib
import re
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from dotenv import load_dotenv

from models import db, Teacher, Student, Class, Attendance, AttendanceRecord

# Load environment variables
load_dotenv()

# Email configuration
EMAIL_HOST = os.getenv('EMAIL_HOST', 'smtp.gmail.com')
EMAIL_PORT = int(os.getenv('EMAIL_PORT', 587))
EMAIL_USER = os.getenv('EMAIL_USER', '')
EMAIL_PASSWORD = os.getenv('EMAIL_PASSWORD', '')
EMAIL_FROM = os.getenv('EMAIL_FROM', '')

def send_email(to_email, subject, message):
    """Send email notification"""
    if not EMAIL_USER or not EMAIL_PASSWORD:
        print(f"Email would be sent to {to_email}: {subject}")
        return False
    
    try:
        msg = MIMEMultipart()
        msg['From'] = EMAIL_FROM or EMAIL_USER
        msg['To'] = to_email
        msg['Subject'] = subject
        
        msg.attach(MIMEText(message, 'html'))
        
        server = smtplib.SMTP(EMAIL_HOST, EMAIL_PORT)
        server.starttls()
        server.login(EMAIL_USER, EMAIL_PASSWORD)
        server.send_message(msg)
        server.quit()
        
        return True
    except Exception as e:
        print(f"Failed to send email: {str(e)}")
        return False

def build_absence_email(student, class_obj, attendance, teacher):
    """Build the standardized absence notification HTML."""
    subject = f"Absence Notification - {class_obj.name} {class_obj.section}"
    message = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
            .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
            .header {{ background-color: #4e73df; color: white; padding: 10px 20px; text-align: center; }}
            .content {{ padding: 20px; border: 1px solid #ddd; border-top: none; }}
            .footer {{ font-size: 12px; text-align: center; margin-top: 20px; color: #777; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h2>Absence Notification</h2>
            </div>
            <div class="content">
                <p>Dear {student.name},</p>
                <p>This is to inform you that you were marked absent in <strong>{class_obj.name} - {class_obj.section}</strong> 
                on <strong>{attendance.date.strftime('%B %d, %Y')}</strong>.</p>
                <p>If you believe this is an error, please contact your teacher, {teacher.name}, at {teacher.email}.</p>
                <p>Thank you,<br>EduGuard AI Attendance System</p>
            </div>
            <div class="footer">
                <p>This is an automated message. Please do not reply to this email.</p>
            </div>
        </div>
    </body>
    </html>
    """
    return subject, message

def normalize_text(value):
    return (value or '').strip()

def is_valid_email(email):
    return bool(re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email or ''))

def is_valid_phone(phone):
    return bool(re.match(r'^[0-9+()\-\s]{7,20}$', phone or ''))

def is_allowed_image_file(filename):
    if not filename:
        return False
    allowed = {'.jpg', '.jpeg', '.png', '.webp'}
    _, ext = os.path.splitext(filename.lower())
    return ext in allowed

def register_routes(app):
    # Authentication routes
    @app.route('/')
    def index():
        if current_user.is_authenticated:
            return redirect(url_for('dashboard'))
        return render_template('index.html')

    @app.route('/login', methods=['GET', 'POST'])
    def login():
        if current_user.is_authenticated:
            return redirect(url_for('dashboard'))
        
        if request.method == 'POST':
            username = normalize_text(request.form.get('username'))
            password = normalize_text(request.form.get('password'))

            if not username or not password:
                flash('Username and password are required.', 'danger')
                return render_template('login.html')
            
            teacher = Teacher.query.filter_by(username=username).first()
            
            if teacher and check_password_hash(teacher.password, password):
                login_user(teacher)
                return redirect(url_for('dashboard'))
            else:
                flash('Login failed. Please check your username and password.', 'danger')
        
        return render_template('login.html')

    @app.route('/register', methods=['GET', 'POST'])
    def register():
        if current_user.is_authenticated:
            return redirect(url_for('dashboard'))
        
        if request.method == 'POST':
            username = normalize_text(request.form.get('username'))
            email = normalize_text(request.form.get('email')).lower()
            password = request.form.get('password') or ''
            name = normalize_text(request.form.get('name'))

            if len(name) < 2 or len(name) > 100:
                flash('Full name must be between 2 and 100 characters.', 'danger')
                return render_template('register.html')

            if len(username) < 3 or len(username) > 80 or not re.match(r'^[A-Za-z0-9_.-]+$', username):
                flash('Username must be 3-80 characters and may include letters, numbers, ., _ and -.', 'danger')
                return render_template('register.html')

            if not is_valid_email(email):
                flash('Please provide a valid email address.', 'danger')
                return render_template('register.html')

            if len(password) < 8:
                flash('Password must be at least 8 characters long.', 'danger')
                return render_template('register.html')
            
            if Teacher.query.filter_by(username=username).first():
                flash('Username already exists.', 'danger')
                return render_template('register.html')
            
            if Teacher.query.filter_by(email=email).first():
                flash('Email already exists.', 'danger')
                return render_template('register.html')
            
            hashed_password = generate_password_hash(password)
            new_teacher = Teacher(username=username, email=email, password=hashed_password, name=name)
            
            db.session.add(new_teacher)
            db.session.commit()
            
            flash('Registration successful! You can now log in.', 'success')
            return redirect(url_for('login'))
        
        return render_template('register.html')

    @app.route('/logout')
    @login_required
    def logout():
        logout_user()
        return redirect(url_for('index'))

    # Dashboard and class management
    @app.route('/dashboard')
    @login_required
    def dashboard():
        classes = Class.query.filter_by(teacher_id=current_user.id).all()
        return render_template('dashboard.html', classes=classes)

    @app.route('/class/new', methods=['GET', 'POST'])
    @login_required
    def new_class():
        if request.method == 'POST':
            name = normalize_text(request.form.get('name'))
            section = normalize_text(request.form.get('section'))

            if len(name) < 2 or len(name) > 100:
                flash('Class name must be between 2 and 100 characters.', 'danger')
                return render_template('new_class.html')

            if len(section) < 1 or len(section) > 20:
                flash('Section is required and must be at most 20 characters.', 'danger')
                return render_template('new_class.html')
            
            new_class = Class(name=name, section=section, teacher_id=current_user.id)
            db.session.add(new_class)
            db.session.commit()
            
            flash('Class created successfully!', 'success')
            return redirect(url_for('dashboard'))
        
        return render_template('new_class.html')

    @app.route('/class/<int:class_id>')
    @login_required
    def view_class(class_id):
        class_obj = Class.query.get_or_404(class_id)
        
        # Ensure the teacher can only view their own classes
        if class_obj.teacher_id != current_user.id:
            flash('You do not have permission to view this class.', 'danger')
            return redirect(url_for('dashboard'))
        
        students = Student.query.filter_by(class_id=class_id).all()
        return render_template('view_class.html', class_obj=class_obj, students=students)

    # Student management
    @app.route('/student/new/<int:class_id>', methods=['GET', 'POST'])
    @login_required
    def new_student(class_id):
        class_obj = Class.query.get_or_404(class_id)
        
        # Ensure the teacher can only add students to their own classes
        if class_obj.teacher_id != current_user.id:
            flash('You do not have permission to add students to this class.', 'danger')
            return redirect(url_for('dashboard'))
        
        if request.method == 'POST':
            student_id = normalize_text(request.form.get('student_id'))
            name = normalize_text(request.form.get('name'))
            email = normalize_text(request.form.get('email')).lower()
            phone = normalize_text(request.form.get('phone'))

            if not student_id or len(student_id) > 50:
                flash('Student ID is required and must be at most 50 characters.', 'danger')
                return render_template('new_student.html', class_obj=class_obj)
            if len(name) < 2 or len(name) > 100:
                flash('Student name must be between 2 and 100 characters.', 'danger')
                return render_template('new_student.html', class_obj=class_obj)
            if not is_valid_email(email):
                flash('Please provide a valid student email address.', 'danger')
                return render_template('new_student.html', class_obj=class_obj)
            if not is_valid_phone(phone):
                flash('Please provide a valid phone number (7-20 chars).', 'danger')
                return render_template('new_student.html', class_obj=class_obj)
            
            # Check if student ID already exists within this class
            if Student.query.filter_by(student_id=student_id, class_id=class_id).first():
                flash('Student ID already exists in this class.', 'danger')
                return render_template('new_student.html', class_obj=class_obj)
            
            # Create new student without face encoding yet
            new_student = Student(
                student_id=student_id,
                name=name,
                email=email,
                phone=phone,
                class_id=class_id
            )
            
            db.session.add(new_student)
            db.session.commit()
            
            # Now process face image if provided
            if 'face_image' in request.files and request.files['face_image'].filename:
                image_file = request.files['face_image']
                if image_file:
                    if not is_allowed_image_file(image_file.filename):
                        flash('Unsupported image format. Please upload JPG, JPEG, PNG, or WEBP.', 'danger')
                        return render_template('new_student.html', class_obj=class_obj)
                    filename = secure_filename(f"student_{new_student.id}_{image_file.filename}")
                    image_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                    image_file.save(image_path)
                    
                    # Process and save face encoding
                    try:
                        image = face_recognition.load_image_file(image_path)
                        face_encodings = face_recognition.face_encodings(image)
                        
                        if face_encodings:
                            # Save the first face encoding found
                            encoding_json = json.dumps(face_encodings[0].tolist())
                            new_student.face_encoding = encoding_json
                            new_student.image_file = filename
                            db.session.commit()
                        else:
                            flash('No face detected in the uploaded image.', 'warning')
                    except Exception as e:
                        flash(f'Error processing face: {str(e)}', 'danger')
            
            flash('Student added successfully!', 'success')
            return redirect(url_for('view_class', class_id=class_id))
        
        return render_template('new_student.html', class_obj=class_obj)

    @app.route('/student/edit/<int:student_id>', methods=['GET', 'POST'])
    @login_required
    def edit_student(student_id):
        student = Student.query.get_or_404(student_id)
        class_obj = Class.query.get_or_404(student.class_id)
        
        # Ensure the teacher can only edit students in their own classes
        if class_obj.teacher_id != current_user.id:
            flash('You do not have permission to edit this student.', 'danger')
            return redirect(url_for('dashboard'))
        
        if request.method == 'POST':
            new_student_id = normalize_text(request.form.get('student_id'))
            new_name = normalize_text(request.form.get('name'))
            new_email = normalize_text(request.form.get('email')).lower()
            new_phone = normalize_text(request.form.get('phone'))

            if not new_student_id or len(new_student_id) > 50:
                flash('Student ID is required and must be at most 50 characters.', 'danger')
                return render_template('edit_student.html', student=student, class_obj=class_obj)
            if len(new_name) < 2 or len(new_name) > 100:
                flash('Student name must be between 2 and 100 characters.', 'danger')
                return render_template('edit_student.html', student=student, class_obj=class_obj)
            if not is_valid_email(new_email):
                flash('Please provide a valid student email address.', 'danger')
                return render_template('edit_student.html', student=student, class_obj=class_obj)
            if not is_valid_phone(new_phone):
                flash('Please provide a valid phone number (7-20 chars).', 'danger')
                return render_template('edit_student.html', student=student, class_obj=class_obj)

            student.name = new_name
            student.email = new_email
            student.phone = new_phone
            
            # If student_id changed, ensure it's unique within the class
            if new_student_id != student.student_id:
                if Student.query.filter_by(student_id=new_student_id, class_id=class_obj.id).first():
                    flash('Student ID already exists in this class.', 'danger')
                    return render_template('edit_student.html', student=student, class_obj=class_obj)
                student.student_id = new_student_id
            
            # Process face image if provided
            if 'face_image' in request.files and request.files['face_image'].filename:
                image_file = request.files['face_image']
                if image_file:
                    if not is_allowed_image_file(image_file.filename):
                        flash('Unsupported image format. Please upload JPG, JPEG, PNG, or WEBP.', 'danger')
                        return render_template('edit_student.html', student=student, class_obj=class_obj)
                    filename = secure_filename(f"student_{student.id}_{image_file.filename}")
                    image_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                    image_file.save(image_path)
                    
                    # Process and save face encoding
                    try:
                        image = face_recognition.load_image_file(image_path)
                        face_encodings = face_recognition.face_encodings(image)
                        
                        if face_encodings:
                            # Save the first face encoding found
                            encoding_json = json.dumps(face_encodings[0].tolist())
                            student.face_encoding = encoding_json
                            student.image_file = filename
                        else:
                            flash('No face detected in the uploaded image.', 'warning')
                    except Exception as e:
                        flash(f'Error processing face: {str(e)}', 'danger')
            
            db.session.commit()
            flash('Student updated successfully!', 'success')
            return redirect(url_for('view_class', class_id=class_obj.id))
        
        return render_template('edit_student.html', student=student, class_obj=class_obj)
    
    @app.route('/student/delete/<int:student_id>', methods=['POST'])
    @login_required
    def delete_student(student_id):
        student = Student.query.get_or_404(student_id)
        class_obj = Class.query.get_or_404(student.class_id)
        
        # Ensure the teacher can only delete students from their own classes
        if class_obj.teacher_id != current_user.id:
            flash('You do not have permission to delete this student.', 'danger')
            return redirect(url_for('dashboard'))
        
        # Delete attendance records for this student
        AttendanceRecord.query.filter_by(student_id=student_id).delete()
        
        # Delete the student
        db.session.delete(student)
        db.session.commit()
        
        flash('Student deleted successfully!', 'success')
        return redirect(url_for('view_class', class_id=class_obj.id))

    # Attendance routes
    @app.route('/attendance/start/<int:class_id>', methods=['GET', 'POST'])
    @login_required
    def start_attendance(class_id):
        class_obj = Class.query.get_or_404(class_id)
        
        # Ensure the teacher can only take attendance for their own classes
        if class_obj.teacher_id != current_user.id:
            flash('You do not have permission to take attendance for this class.', 'danger')
            return redirect(url_for('dashboard'))
        
        students = Student.query.filter_by(class_id=class_id).all()
        
        if request.method == 'POST':
            attendance_date_raw = normalize_text(request.form.get('attendance_date'))
            if not attendance_date_raw:
                flash('Attendance date is required.', 'danger')
                return render_template('start_attendance.html', class_obj=class_obj, students=students)

            try:
                attendance_date = datetime.strptime(attendance_date_raw, '%Y-%m-%d').date()
            except ValueError:
                flash('Invalid attendance date format.', 'danger')
                return render_template('start_attendance.html', class_obj=class_obj, students=students)
            
            # Check if attendance for this date and class already exists
            existing_attendance = Attendance.query.filter_by(
                date=attendance_date,
                class_id=class_id
            ).first()
            
            if existing_attendance:
                flash('Attendance for this date already exists.', 'warning')
                return redirect(url_for('view_attendance', attendance_id=existing_attendance.id))
            
            # Create new attendance record
            new_attendance = Attendance(
                date=attendance_date,
                class_id=class_id,
                teacher_id=current_user.id
            )
            
            db.session.add(new_attendance)
            db.session.commit()
            
            # Create attendance records for all students, initially marked as absent
            for student in students:
                record = AttendanceRecord(
                    attendance_id=new_attendance.id,
                    student_id=student.id,
                    status=False
                )
                db.session.add(record)
            
            db.session.commit()
            
            return redirect(url_for('take_attendance', attendance_id=new_attendance.id))
        
        return render_template('start_attendance.html', class_obj=class_obj, students=students)

    @app.route('/attendance/take/<int:attendance_id>')
    @login_required
    def take_attendance(attendance_id):
        attendance = Attendance.query.get_or_404(attendance_id)
        class_obj = Class.query.get_or_404(attendance.class_id)
        
        # Ensure the teacher can only view their own attendance records
        if attendance.teacher_id != current_user.id:
            flash('You do not have permission to view this attendance record.', 'danger')
            return redirect(url_for('dashboard'))
        
        students = Student.query.filter_by(class_id=class_obj.id).all()
        
        return render_template('take_attendance.html', 
                              attendance=attendance, 
                              class_obj=class_obj, 
                              students=students)

    @app.route('/api/mark_attendance', methods=['POST'])
    @login_required
    def mark_attendance():
        data = request.get_json(silent=True) or {}

        try:
            student_id = int(data.get('student_id'))
            attendance_id = int(data.get('attendance_id'))
        except (TypeError, ValueError):
            return jsonify({'success': False, 'error': 'Invalid student or attendance id'}), 400

        status = bool(data.get('status', True))  # Default to present

        attendance = Attendance.query.get(attendance_id)
        if not attendance:
            return jsonify({'success': False, 'error': 'Attendance not found'}), 404

        # Ensure a teacher can only edit their own attendance sessions
        if attendance.teacher_id != current_user.id:
            return jsonify({'success': False, 'error': 'Permission denied'}), 403

        record = AttendanceRecord.query.filter_by(
            attendance_id=attendance_id,
            student_id=student_id
        ).first()
        if not record:
            return jsonify({'success': False, 'error': 'Record not found'}), 404

        # Keep this endpoint fast: only persist status here.
        # Notifications are sent via /attendance/send_notifications.
        record.status = status
        db.session.commit()

        return jsonify({
            'success': True,
            'status': status,
            'notification_sent': bool(record.notification_sent)
        })

    @app.route('/attendance/view/<int:attendance_id>')
    @login_required
    def view_attendance(attendance_id):
        attendance = Attendance.query.get_or_404(attendance_id)
        
        # Ensure the teacher can only view their own attendance records
        if attendance.teacher_id != current_user.id:
            flash('You do not have permission to view this attendance record.', 'danger')
            return redirect(url_for('dashboard'))
        
        class_obj = Class.query.get(attendance.class_id)
        records = AttendanceRecord.query.filter_by(attendance_id=attendance_id).all()
        
        # Get student details for each record
        attendance_data = []
        for record in records:
            student = Student.query.get(record.student_id)
            if student:
                attendance_data.append({
                    'record': record,
                    'student': student
                })
        
        present_count = sum(1 for record in records if record.status)
        absent_count = len(records) - present_count
        
        return render_template('view_attendance.html', 
                              attendance=attendance, 
                              class_obj=class_obj,
                              attendance_data=attendance_data,
                              present_count=present_count,
                              absent_count=absent_count)

    @app.route('/attendance/export/pdf/<int:attendance_id>')
    @login_required
    def export_attendance_pdf(attendance_id):
        attendance = Attendance.query.get_or_404(attendance_id)
        
        # Ensure the teacher can only export their own attendance records
        if attendance.teacher_id != current_user.id:
            flash('You do not have permission to export this attendance record.', 'danger')
            return redirect(url_for('dashboard'))
        
        class_obj = Class.query.get(attendance.class_id)
        records = AttendanceRecord.query.filter_by(attendance_id=attendance_id).all()
        
        # Get student details for each record
        attendance_data = []
        for record in records:
            student = Student.query.get(record.student_id)
            if student:
                attendance_data.append({
                    'record': record,
                    'student': student
                })
        
        present_count = sum(1 for record in records if record.status)
        absent_count = len(records) - present_count
        
        # Create a PDF using ReportLab
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        
        # Add title
        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, height - 50, f"Attendance Report: {class_obj.name} - {class_obj.section}")
        
        # Add date
        p.setFont("Helvetica", 12)
        p.drawString(50, height - 80, f"Date: {attendance.date.strftime('%B %d, %Y')}")
        
        # Add summary
        p.drawString(50, height - 110, f"Present: {present_count} | Absent: {absent_count}")
        
        # Add table headers
        p.setFont("Helvetica-Bold", 10)
        p.drawString(50, height - 150, "Student ID")
        p.drawString(150, height - 150, "Name")
        p.drawString(300, height - 150, "Status")
        p.drawString(400, height - 150, "Notification")
        
        # Add horizontal line
        p.line(50, height - 160, 550, height - 160)
        
        # Add table data
        y_position = height - 180
        p.setFont("Helvetica", 10)
        
        for item in attendance_data:
            p.drawString(50, y_position, item['student'].student_id)
            p.drawString(150, y_position, item['student'].name)
            p.drawString(300, y_position, "Present" if item['record'].status else "Absent")
            
            if not item['record'].status:
                notification_status = "Sent" if item['record'].notification_sent else "Pending"
            else:
                notification_status = "N/A"
            
            p.drawString(400, y_position, notification_status)
            
            y_position -= 20
            
            # Add a new page if we run out of space
            if y_position < 50:
                p.showPage()
                p.setFont("Helvetica-Bold", 10)
                p.drawString(50, height - 50, "Student ID")
                p.drawString(150, height - 50, "Name")
                p.drawString(300, height - 50, "Status")
                p.drawString(400, height - 50, "Notification")
                p.line(50, height - 60, 550, height - 60)
                p.setFont("Helvetica", 10)
                y_position = height - 80
        
        p.showPage()
        p.save()
        
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"attendance_{class_obj.name}_{attendance.date.strftime('%Y-%m-%d')}.pdf",
            mimetype='application/pdf'
        )

    @app.route('/attendance/export/excel/<int:attendance_id>')
    @login_required
    def export_attendance_excel(attendance_id):
        attendance = Attendance.query.get_or_404(attendance_id)
        
        # Ensure the teacher can only export their own attendance records
        if attendance.teacher_id != current_user.id:
            flash('You do not have permission to export this attendance record.', 'danger')
            return redirect(url_for('dashboard'))
        
        class_obj = Class.query.get(attendance.class_id)
        records = AttendanceRecord.query.filter_by(attendance_id=attendance_id).all()
        
        # Get student details for each record
        attendance_data = []
        for record in records:
            student = Student.query.get(record.student_id)
            if student:
                attendance_data.append({
                    'student_id': student.student_id,
                    'name': student.name,
                    'email': student.email,
                    'phone': student.phone,
                    'status': 'Present' if record.status else 'Absent',
                    'notification': 'Sent' if not record.status and record.notification_sent else 
                                   'Pending' if not record.status and not record.notification_sent else 'N/A'
                })
        
        # Create Excel file
        output = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Attendance"
        
        # Add headers
        headers = ['Student ID', 'Name', 'Email', 'Phone', 'Status', 'Notification']
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for row_num, item in enumerate(attendance_data, 2):
            worksheet.cell(row=row_num, column=1).value = item['student_id']
            worksheet.cell(row=row_num, column=2).value = item['name']
            worksheet.cell(row=row_num, column=3).value = item['email']
            worksheet.cell(row=row_num, column=4).value = item['phone']
            worksheet.cell(row=row_num, column=5).value = item['status']
            worksheet.cell(row=row_num, column=6).value = item['notification']
            
            # Color coding for status
            if item['status'] == 'Present':
                worksheet.cell(row=row_num, column=5).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                worksheet.cell(row=row_num, column=5).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Add summary
        summary_row = len(attendance_data) + 3
        worksheet.cell(row=summary_row, column=1).value = "Summary"
        worksheet.cell(row=summary_row, column=1).font = Font(bold=True)
        
        present_count = sum(1 for item in attendance_data if item['status'] == 'Present')
        absent_count = len(attendance_data) - present_count
        
        worksheet.cell(row=summary_row + 1, column=1).value = "Present:"
        worksheet.cell(row=summary_row + 1, column=2).value = present_count
        
        worksheet.cell(row=summary_row + 2, column=1).value = "Absent:"
        worksheet.cell(row=summary_row + 2, column=2).value = absent_count
        
        # Add class and date info
        worksheet.cell(row=summary_row + 4, column=1).value = "Class:"
        worksheet.cell(row=summary_row + 4, column=2).value = f"{class_obj.name} - {class_obj.section}"
        
        worksheet.cell(row=summary_row + 5, column=1).value = "Date:"
        worksheet.cell(row=summary_row + 5, column=2).value = attendance.date.strftime('%B %d, %Y')
        
        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        workbook.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=f"attendance_{class_obj.name}_{attendance.date.strftime('%Y-%m-%d')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    @app.route('/class/export/excel/<int:class_id>')
    @login_required
    def export_students_excel(class_id):
        class_obj = Class.query.get_or_404(class_id)
        
        # Ensure the teacher can only export their own class data
        if class_obj.teacher_id != current_user.id:
            flash('You do not have permission to export this class data.', 'danger')
            return redirect(url_for('dashboard'))
        
        students = Student.query.filter_by(class_id=class_id).all()
        
        # Create Excel file
        output = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Students"
        
        # Add headers
        headers = ['Student ID', 'Name', 'Email', 'Phone', 'Face Recognition']
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for row_num, student in enumerate(students, 2):
            worksheet.cell(row=row_num, column=1).value = student.student_id
            worksheet.cell(row=row_num, column=2).value = student.name
            worksheet.cell(row=row_num, column=3).value = student.email
            worksheet.cell(row=row_num, column=4).value = student.phone
            worksheet.cell(row=row_num, column=5).value = "Registered" if student.face_encoding else "Not Registered"
            
            # Color coding for face recognition status
            if student.face_encoding:
                worksheet.cell(row=row_num, column=5).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                worksheet.cell(row=row_num, column=5).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # Add class info
        summary_row = len(students) + 3
        worksheet.cell(row=summary_row, column=1).value = "Class:"
        worksheet.cell(row=summary_row, column=2).value = f"{class_obj.name} - {class_obj.section}"
        
        worksheet.cell(row=summary_row + 1, column=1).value = "Total Students:"
        worksheet.cell(row=summary_row + 1, column=2).value = len(students)
        
        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        workbook.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=f"students_{class_obj.name}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    @app.route('/class/export/pdf/<int:class_id>')
    @login_required
    def export_students_pdf(class_id):
        class_obj = Class.query.get_or_404(class_id)
        
        # Ensure the teacher can only export their own class data
        if class_obj.teacher_id != current_user.id:
            flash('You do not have permission to export this class data.', 'danger')
            return redirect(url_for('dashboard'))
        
        students = Student.query.filter_by(class_id=class_id).all()
        
        # Create a PDF using ReportLab
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        
        # Add title
        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, height - 50, f"Student List: {class_obj.name} - {class_obj.section}")
        
        # Add table headers
        p.setFont("Helvetica-Bold", 10)
        p.drawString(50, height - 100, "Student ID")
        p.drawString(150, height - 100, "Name")
        p.drawString(300, height - 100, "Email")
        p.drawString(450, height - 100, "Phone")
        
        # Add horizontal line
        p.line(50, height - 110, 550, height - 110)
        
        # Add table data
        y_position = height - 130
        p.setFont("Helvetica", 10)
        
        for student in students:
            p.drawString(50, y_position, student.student_id)
            p.drawString(150, y_position, student.name)
            p.drawString(300, y_position, student.email)
            p.drawString(450, y_position, student.phone)
            
            y_position -= 20
            
            # Add a new page if we run out of space
            if y_position < 50:
                p.showPage()
                p.setFont("Helvetica-Bold", 10)
                p.drawString(50, height - 50, "Student ID")
                p.drawString(150, height - 50, "Name")
                p.drawString(300, height - 50, "Email")
                p.drawString(450, height - 50, "Phone")
                p.line(50, height - 60, 550, height - 60)
                p.setFont("Helvetica", 10)
                y_position = height - 80
        
        p.showPage()
        p.save()
        
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"students_{class_obj.name}.pdf",
            mimetype='application/pdf'
        )

    @app.route('/attendance/send_notifications/<int:attendance_id>', methods=['POST'])
    @login_required
    def send_notifications(attendance_id):
        attendance = Attendance.query.get_or_404(attendance_id)
        
        # Ensure the teacher can only send notifications for their own attendance records
        if attendance.teacher_id != current_user.id:
            flash('You do not have permission to send notifications for this attendance record.', 'danger')
            return redirect(url_for('dashboard'))
        
        class_obj = Class.query.get(attendance.class_id)
        records = AttendanceRecord.query.filter_by(
            attendance_id=attendance_id,
            status=False,  # Only absent students
            notification_sent=False  # Only those who haven't received notifications yet
        ).all()
        
        sent_count = 0
        failed_count = 0
        teacher = Teacher.query.get(attendance.teacher_id)

        for record in records:
            student = Student.query.get(record.student_id)
            if not student:
                continue

            subject, message = build_absence_email(student, class_obj, attendance, teacher)
            email_sent = send_email(student.email, subject, message)

            if email_sent:
                record.notification_sent = True
                sent_count += 1
            else:
                failed_count += 1

        db.session.commit()
        
        if sent_count > 0:
            flash(f'Sent {sent_count} absence notifications successfully!', 'success')
        if failed_count > 0:
            flash(f'Failed to send {failed_count} notifications. Check email settings.', 'warning')
        if sent_count == 0 and failed_count == 0:
            flash('No pending notifications to send.', 'info')
        
        return redirect(url_for('view_attendance', attendance_id=attendance_id))

    @app.route('/api/start_face_detection/<int:attendance_id>', methods=['POST'])
    @login_required
    def start_face_detection(attendance_id):
        """Start face detection in a separate window"""
        from app import start_face_detection
        
        attendance = Attendance.query.get_or_404(attendance_id)
        
        # Ensure the teacher can only view their own attendance records
        if attendance.teacher_id != current_user.id:
            return jsonify({
                'success': False, 
                'message': 'You do not have permission to take attendance for this class.'
            })
        
        # Create a temp file for results
        import tempfile
        result_file = os.path.join(tempfile.gettempdir(), f"eduguard_results_{attendance_id}.json")
        
        # Resolve the single source of truth DB path: Flask instance folder DB
        configured_path = app.config['SQLALCHEMY_DATABASE_URI'].replace('sqlite:///', '')
        if os.path.isabs(configured_path):
            db_path = configured_path
        else:
            # Flask stores relative SQLite DBs under instance/ by default
            db_path = os.path.join(app.instance_path, configured_path)

        print(f"Starting face detection for attendance ID: {attendance_id}")
        print(f"Using database: {db_path}")
        
        # Start face detection and explicitly pass the attendance_id
        result = start_face_detection(db_path, result_file, attendance_id)
        
        return jsonify(result)
    
    @app.route('/api/check_face_detection_results/<int:attendance_id>', methods=['GET'])
    @login_required
    def check_face_detection_results(attendance_id):
        """Check if there are any results from the face detection process"""
        attendance = Attendance.query.get_or_404(attendance_id)
        if attendance.teacher_id != current_user.id:
            return jsonify({
                'success': False,
                'message': 'Permission denied',
                'marked_students': []
            }), 403

        import tempfile
        result_file = os.path.join(tempfile.gettempdir(), f"eduguard_results_{attendance_id}.json")
        
        if not os.path.exists(result_file):
            return jsonify({
                'success': True,
                'has_results': False,
                'marked_students': []
            })
        
        try:
            with open(result_file, 'r') as f:
                results = json.load(f)
            
            # Convert student IDs to names for display
            marked_students = []
            for student_id in results.get('marked_students', []):
                student = Student.query.filter_by(
                    student_id=student_id,
                    class_id=attendance.class_id
                ).first()
                if student:
                    marked_students.append({
                        'id': student.id,
                        'name': student.name,
                        'student_id': student_id
                    })
            
            return jsonify({
                'success': True,
                'has_results': True,
                'marked_students': marked_students
            })
        except Exception as e:
            return jsonify({
                'success': False,
                'message': f"Error reading results: {str(e)}",
                'marked_students': []
            }) 
